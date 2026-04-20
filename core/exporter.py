"""
Excel exporter.
Two modes:
  1. export_excel()        — modifies original file in-place (eMAG Offer ch. N format)
  2. export_model_format() — builds new file from scratch in offers_model_import format
"""
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from collections import defaultdict


# Colour palette (8-char ARGB so openpyxl round-trips correctly)
C_GREEN_BG  = "FFC6EFCE"; C_GREEN_FG  = "FF276221"
C_BLUE_BG   = "FFDDEBF7"; C_BLUE_FG   = "FF1F4E79"
C_ORANGE_BG = "FFFCE4D6"; C_ORANGE_FG = "FF843C0C"
C_RED_BG    = "FFFFC7CE"; C_RED_FG    = "FF9C0006"
C_YELLOW_BG = "FFFFEB9C"; C_YELLOW_FG = "FF9C6500"

def _fill(bg): return PatternFill("solid", start_color=bg, end_color=bg)
def _font(fg, bold=False): return Font(color=fg, name="Calibri", bold=bold)


def export_model_format(
    original_file,
    results: list[dict],
    products: list[dict],
) -> bytes:
    """
    Builds a new Excel file in offers_model_import format:
    Internal ID | Name | Brand | Description | Category |
    Characteristic Name | Characteristic Value |
    Characteristic Name.1 | Characteristic Value.1 | ...

    Works regardless of what columns the original file had.
    """
    # Determine max number of characteristic pairs needed
    max_chars = max(
        (len({**p.get("existing_chars", {}), **r.get("new_chars", {}),
              **{ch: "" for ch in r.get("missing_mandatory", [])
                 if ch not in {**p.get("existing_chars", {}), **r.get("new_chars", {})}}})
         for p, r in zip(products, results)),
        default=1
    )
    max_chars = max(max_chars, 1)

    # Build header
    headers = ["Internal ID", "Name", "Brand", "Description", "Category"]
    for i in range(max_chars):
        suffix = "" if i == 0 else f".{i}"
        headers += [f"Characteristic Name{suffix}", f"Characteristic Value{suffix}"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name="Calibri")

    for row_num, (prod, result) in enumerate(zip(products, results), start=2):
        action     = result.get("action", "skip")
        new_cat    = result.get("new_category") or prod.get("category") or ""
        new_chars  = result.get("new_chars", {})
        existing   = dict(prod.get("existing_chars") or {})
        cleared    = set(result.get("cleared", []))
        needs_manual = result.get("needs_manual", False)

        # Merge existing (minus cleared) + new
        all_chars = {k: v for k, v in existing.items() if k not in cleared and v}
        all_chars.update(new_chars)
        # Inject missing_mandatory as empty entries (for red coloring)
        for ch in result.get("missing_mandatory", []):
            if ch not in all_chars:
                all_chars[ch] = ""

        # Base fields
        ws.cell(row=row_num, column=1, value=prod.get("id"))
        ws.cell(row=row_num, column=2, value=prod.get("name"))
        ws.cell(row=row_num, column=3, value=prod.get("brand"))
        ws.cell(row=row_num, column=4, value=prod.get("description"))

        # Category cell
        cat_cell = ws.cell(row=row_num, column=5, value=new_cat)
        if action == "cat_assigned":
            cat_cell.fill = _fill(C_BLUE_BG)
            cat_cell.font = _font(C_BLUE_FG)
        elif action == "cat_corrected":
            cat_cell.fill = _fill(C_ORANGE_BG)
            cat_cell.font = _font(C_ORANGE_FG)

        # Characteristic pairs
        missing_mandatory = result.get("missing_mandatory", [])
        for i, (char_name, char_val) in enumerate(all_chars.items()):
            col_name = 6 + i * 2
            col_val  = 7 + i * 2
            is_new   = char_name in new_chars
            is_cleared = char_name in cleared
            is_missing_mandatory = char_name in missing_mandatory

            name_cell = ws.cell(row=row_num, column=col_name, value=char_name)
            val_cell  = ws.cell(row=row_num, column=col_val,  value=char_val if char_val else None)

            if is_new:
                for c in (name_cell, val_cell):
                    c.fill = _fill(C_GREEN_BG)
                    c.font = _font(C_GREEN_FG)
            elif is_cleared:
                val_cell.value = None
                val_cell.fill  = _fill(C_RED_BG)
                val_cell.font  = _font(C_RED_FG)
            elif is_missing_mandatory:
                for c in (name_cell, val_cell):
                    c.fill = _fill(C_RED_BG)
                    c.font = _font(C_RED_FG)

        # Mark manual review row
        if needs_manual:
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", ""):
                    cell.fill = _fill(C_YELLOW_BG)

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_excel(
    original_file,
    results: list[dict],   # one dict per product row (same order as file)
    char_pairs: list,
) -> bytes:
    """
    Re-open the original workbook, apply all changes with colour coding,
    return bytes of the modified workbook.
    """
    wb = openpyxl.load_workbook(original_file)
    sheet_name = next((s for s in wb.sheetnames if "export" in s.lower()), wb.sheetnames[0])
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]

    # Build col index: header value -> 1-based col number
    col_idx = {str(h).strip() if h else "": i + 1 for i, h in enumerate(headers)}

    # Style header row
    for cell in ws[1]:
        cell.font = Font(bold=True, name="Calibri")

    for row_num, result in enumerate(results, start=2):
        if row_num > ws.max_row:
            break

        action = result.get("action", "")

        # ── Category cell (col F = index 6) ────────────────────────────────
        cat_col = col_idx.get("Categorie", 6)
        cat_cell = ws.cell(row=row_num, column=cat_col)
        if action == "cat_assigned":
            cat_cell.value = result["new_category"]
            cat_cell.fill  = _fill(C_BLUE_BG)
            cat_cell.font  = _font(C_BLUE_FG)
        elif action == "cat_corrected":
            cat_cell.value = result["new_category"]
            cat_cell.fill  = _fill(C_ORANGE_BG)
            cat_cell.font  = _font(C_ORANGE_FG)

        # ── Clear invalid existing values ────────────────────────────────────
        for char_name in result.get("cleared", []):
            for name_col, val_col in char_pairs:
                name_c = col_idx.get(name_col)
                val_c  = col_idx.get(val_col)
                if name_c and ws.cell(row=row_num, column=name_c).value == char_name:
                    cell = ws.cell(row=row_num, column=val_c)
                    cell.value = None
                    cell.fill  = _fill(C_RED_BG)
                    cell.font  = _font(C_RED_FG)
                    break

        # ── Add new characteristic values ────────────────────────────────────
        new_chars = result.get("new_chars", {})
        for char_name, char_val in new_chars.items():
            # Find first empty slot
            for name_col, val_col in char_pairs:
                name_c = col_idx.get(name_col)
                val_c  = col_idx.get(val_col)
                if name_c and ws.cell(row=row_num, column=name_c).value is None:
                    ws.cell(row=row_num, column=name_c).value = char_name
                    ws.cell(row=row_num, column=name_c).fill  = _fill(C_GREEN_BG)
                    ws.cell(row=row_num, column=name_c).font  = _font(C_GREEN_FG)
                    ws.cell(row=row_num, column=val_c).value  = char_val
                    ws.cell(row=row_num, column=val_c).fill   = _fill(C_GREEN_BG)
                    ws.cell(row=row_num, column=val_c).font   = _font(C_GREEN_FG)
                    break

        # ── Add missing mandatory chars (red) into first free slot ───────────
        for char_name in result.get("missing_mandatory", []):
            if char_name in new_chars:
                continue  # already written green above
            for name_col, val_col in char_pairs:
                name_c = col_idx.get(name_col)
                val_c  = col_idx.get(val_col)
                if name_c and ws.cell(row=row_num, column=name_c).value is None:
                    ws.cell(row=row_num, column=name_c).value = char_name
                    ws.cell(row=row_num, column=name_c).fill  = _fill(C_RED_BG)
                    ws.cell(row=row_num, column=name_c).font  = _font(C_RED_FG)
                    ws.cell(row=row_num, column=val_c).value  = None
                    ws.cell(row=row_num, column=val_c).fill   = _fill(C_RED_BG)
                    ws.cell(row=row_num, column=val_c).font   = _font(C_RED_FG)
                    break

        # ── Mark rows still needing manual review ────────────────────────────
        if result.get("needs_manual"):
            err_col = col_idx.get("Eroare ofertă", col_idx.get("Eroare oferta", 4))
            cell = ws.cell(row=row_num, column=err_col)
            cell.fill = _fill(C_YELLOW_BG)
            cell.font = _font(C_YELLOW_FG)

    # Return as bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
