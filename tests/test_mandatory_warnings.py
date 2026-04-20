"""Tests for Task 1A: warning when mandatory char has no valid_values."""
import pytest
from unittest.mock import MagicMock


def _make_data_mock(valid_values_map: dict):
    """Create a MarketplaceData mock where valid_values returns per-char sets."""
    data = MagicMock()
    data.valid_values.side_effect = lambda cat_id, ch: valid_values_map.get(ch, set())
    return data


def test_warning_when_valid_values_empty(caplog):
    """ATENȚIE warning fires when valid_values is empty for a mandatory missing char."""
    import logging
    from core.processor import _warn_missing_mandatory_no_values

    data = _make_data_mock({"Marime:": set()})

    with caplog.at_level(logging.WARNING, logger="marketplace.processor"):
        _warn_missing_mandatory_no_values(data, cat_id=123, still_missing=["Marime:"])

    assert any("[ATENȚIE]" in r.message for r in caplog.records)
    assert any("Marime:" in r.message for r in caplog.records)


def test_no_warning_when_valid_values_populated(caplog):
    """ATENȚIE warning does NOT fire when valid_values is non-empty."""
    import logging
    from core.processor import _warn_missing_mandatory_no_values

    data = _make_data_mock({"Marime:": {"S", "M", "L"}})

    with caplog.at_level(logging.WARNING, logger="marketplace.processor"):
        _warn_missing_mandatory_no_values(data, cat_id=123, still_missing=["Marime:"])

    assert not any("[ATENȚIE]" in r.message for r in caplog.records)


def test_export_model_format_red_for_missing_mandatory():
    """missing_mandatory chars not in new_chars get red fill in export_model_format."""
    import io
    import openpyxl
    from core.exporter import export_model_format, C_RED_BG

    result = {
        "action": "skip",
        "new_chars": {},
        "existing_chars": {},
        "cleared": [],
        "needs_manual": False,
        "missing_mandatory": ["Marime:"],
    }
    product = {"id": "1", "name": "Test", "brand": "X", "description": "", "category": "Cat"}

    raw = export_model_format(None, [result], [product])
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb.active

    found_red = False
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value == "Marime:":
                assert cell.fill.fgColor.rgb.upper() == C_RED_BG.upper(), \
                    f"Expected red fill, got {cell.fill.fgColor.rgb}"
                found_red = True
    assert found_red, "Marime: cell not found in worksheet"


def test_export_model_format_green_not_overridden_by_missing_mandatory():
    """A char in both new_chars and missing_mandatory stays GREEN."""
    import io
    import openpyxl
    from core.exporter import export_model_format, C_GREEN_BG

    result = {
        "action": "skip",
        "new_chars": {"Marime:": "M"},
        "existing_chars": {},
        "cleared": [],
        "needs_manual": False,
        "missing_mandatory": ["Marime:"],
    }
    product = {"id": "1", "name": "Test", "brand": "X", "description": "", "category": "Cat"}

    raw = export_model_format(None, [result], [product])
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value == "Marime:":
                assert cell.fill.fgColor.rgb.upper() == C_GREEN_BG.upper(), \
                    f"Expected green fill, got {cell.fill.fgColor.rgb}"


def test_export_model_format_no_missing_mandatory_key():
    """result without missing_mandatory key -> no red cells added."""
    import io
    import openpyxl
    from core.exporter import export_model_format, C_RED_BG

    result = {
        "action": "skip",
        "new_chars": {},
        "existing_chars": {},
        "cleared": [],
        "needs_manual": False,
    }
    product = {"id": "1", "name": "Test", "brand": "X", "description": "", "category": "Cat"}

    raw = export_model_format(None, [result], [product])
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            rgb = cell.fill.fgColor.rgb.upper() if cell.fill and cell.fill.fgColor else ""
            assert rgb != C_RED_BG.upper(), f"Unexpected red cell at {cell.coordinate}"


def test_export_excel_red_for_missing_mandatory():
    """missing_mandatory chars not in new_chars get written in red in export_excel."""
    import io
    import openpyxl
    from core.exporter import export_excel, C_RED_BG

    # Build a minimal workbook with header + 1 data row + char pair columns
    wb_in = openpyxl.Workbook()
    ws_in = wb_in.active
    ws_in.title = "export"
    ws_in.append(["Categorie", "Characteristic Name", "Characteristic Value"])
    ws_in.append(["Pantofi", None, None])

    buf = io.BytesIO()
    wb_in.save(buf)
    buf.seek(0)

    result = {
        "action": "skip",
        "new_chars": {},
        "cleared": [],
        "needs_manual": False,
        "missing_mandatory": ["Marime:"],
    }
    char_pairs = [("Characteristic Name", "Characteristic Value")]

    raw = export_excel(buf, [result], char_pairs)
    wb_out = openpyxl.load_workbook(io.BytesIO(raw))
    ws_out = wb_out.active

    name_cell = ws_out.cell(row=2, column=2)
    assert name_cell.value == "Marime:"
    assert name_cell.fill.fgColor.rgb.upper() == C_RED_BG.upper()

    val_cell = ws_out.cell(row=2, column=3)
    assert val_cell.fill.fgColor.rgb.upper() == C_RED_BG.upper()
