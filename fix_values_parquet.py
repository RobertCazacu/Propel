"""
Script de conversie one-time pentru characteristic_values.xlsx → values.parquet

Problema: fisierul de valori nu are category_id si characteristic_name direct.
Solutia: join cu characteristics.parquet pentru a obtine aceste coloane.

Rulare: python fix_values_parquet.py
"""
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

VALUES_XLSX  = Path("C:/Users/manue/Desktop/characteristic_values.xlsx")
PARQUET_DIR  = Path("data/eMAG_Romania")
CHARS_PARQ   = PARQUET_DIR / "characteristics.parquet"
OUT_PARQ     = PARQUET_DIR / "values.parquet"

BATCH_SIZE = 50_000  # randuri procesate odata

def main():
    print("Citesc characteristics.parquet...")
    chars = pd.read_parquet(CHARS_PARQ, columns=["id", "category_id", "name"])
    chars = chars.rename(columns={"id": "char_id", "name": "characteristic_name"})
    chars["char_id"] = pd.to_numeric(chars["char_id"], errors="coerce")
    print(f"  {len(chars)} caracteristici incarcate.")

    print(f"Deschid {VALUES_XLSX} in mod streaming (read-only)...")
    wb = load_workbook(VALUES_XLSX, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    print(f"  Coloane gasit in fisier: {list(header)}")

    # Detecteaza indecsi coloane
    header_lower = [str(h).lower().strip() if h else "" for h in header]
    def col_idx(*names):
        for n in names:
            if n in header_lower:
                return header_lower.index(n)
        return None

    idx_char_id = col_idx("emag_characteristic_id", "characteristic_id", "char_id")
    idx_value   = col_idx("value", "value_ro", "valoare")

    if idx_char_id is None or idx_value is None:
        print(f"EROARE: nu am gasit coloanele necesare!")
        print(f"  emag_characteristic_id index: {idx_char_id}")
        print(f"  value index: {idx_value}")
        return

    print(f"  Coloana char_id: index {idx_char_id} ('{header[idx_char_id]}')")
    print(f"  Coloana value:   index {idx_value} ('{header[idx_value]}')")

    all_chunks = []
    batch = []
    total_rows = 0
    total_valid = 0

    print("Procesez randuri...")
    for row in rows_iter:
        char_id_raw = row[idx_char_id]
        value_raw   = row[idx_value]

        if char_id_raw is None or value_raw is None:
            continue

        batch.append({
            "characteristic_id": char_id_raw,
            "value": str(value_raw).strip(),
        })
        total_rows += 1

        if len(batch) >= BATCH_SIZE:
            chunk_df = pd.DataFrame(batch)
            chunk_df["characteristic_id"] = pd.to_numeric(chunk_df["characteristic_id"], errors="coerce")
            merged = chunk_df.merge(chars, left_on="characteristic_id", right_on="char_id", how="left")
            merged = merged.dropna(subset=["category_id"])
            merged = merged[["category_id", "characteristic_id", "characteristic_name", "value"]]
            all_chunks.append(merged)
            total_valid += len(merged)
            batch = []
            print(f"  Procesat {total_rows:,} randuri | valori valide: {total_valid:,}")

    # Ultimul batch
    if batch:
        chunk_df = pd.DataFrame(batch)
        chunk_df["characteristic_id"] = pd.to_numeric(chunk_df["characteristic_id"], errors="coerce")
        merged = chunk_df.merge(chars, left_on="characteristic_id", right_on="char_id", how="left")
        merged = merged.dropna(subset=["category_id"])
        merged = merged[["category_id", "characteristic_id", "characteristic_name", "value"]]
        all_chunks.append(merged)
        total_valid += len(merged)

    wb.close()

    print(f"\nTotal randuri citite: {total_rows:,}")
    print(f"Total valori cu category_id: {total_valid:,}")

    if not all_chunks:
        print("EROARE: niciun rand valid gasit!")
        return

    print("Concatenez si salvez parquet...")
    final_df = pd.concat(all_chunks, ignore_index=True)
    final_df.to_parquet(OUT_PARQ, index=False)
    print(f"Salvat: {OUT_PARQ}")
    print(f"Shape final: {final_df.shape}")
    print(f"\nEXAMPLE (primele 5 randuri):")
    print(final_df.head())
    print("\nGata! Reporneste aplicatia Streamlit.")

if __name__ == "__main__":
    main()
